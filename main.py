###########################################################
# Digital register for School Number 4                    #
# Created on 24/02/2020                                   #
# Creator: Max Shevtsov(Nimchenko), 2019-2020 - 11 form   #
###########################################################

from PyQt5.QtWidgets import QApplication, QLabel, QWidget, QVBoxLayout, QLineEdit, \
    QPushButton, QMainWindow, QHBoxLayout, QMessageBox, \
    QComboBox, QBoxLayout, QTableWidget, QAbstractItemView, \
    QTableWidgetItem, QTextEdit, QGridLayout
from PyQt5.QtGui import QPixmap, QColor, QIcon
from PyQt5.QtCore import QSize
import openpyxl
import datetime
import pyodbc
import sys
import os

# Reading a cwd.
cwd = str(os.getcwd())

# Connect data base.
try:
    cnxn = pyodbc.connect(r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"
                          r"DBQ=" + cwd + "/files/data.accdb")
    cursor = cnxn.cursor()
except:
    pass

# Information window.


class InfoWindow(QWidget):
    def __init__(self):
        QWidget.__init__(self)

        # Window.
        self.setFixedSize(500, 550)
        self.setWindowTitle('Інформація')

        # Layout.
        self.mainLayout = QVBoxLayout(self)

        # Main info label.
        self.infoLabel1 = QLabel()
        self.infoLabel2 = QLabel()
        self.linkLabel = QLabel(
            '<a href=\"https://lzosh4.at.ua/\">https://lzosh4.at.ua/</a>')
        self.linkLabel.setOpenExternalLinks(True)
        self.infoLabel1.setText('---Загальна інформація---\n'
                                'Електронний журнал ЛЗЗСО I-III ст. №4 v.1.0, 2019 рік\n'
                                'Автор та розробник: Німченко Максим\n'
                                'Адміністратор: Ященко Тамара\n'
                                'Зворотний зв\'язок:\n'
                                'Instagram: @iinteeel\n'
                                'Email: maks.nimchenko@gmail.com\n'
                                'Офіційний сайт:')
        self.infoLabel2.setText('---Інструкція по користуванню електронним щоденником---\n'
                                'Для зручності роботи з програмою можна використати такі гарячі клавіші:\n'
                                '- Enter - увійти в обліковий запис;\n'
                                '- Ctrl + Q - вийти з облікового запису;\n'
                                '- Ctrl + R - оновити сторінку;\n'
                                '- Ctrl + F - надіслати відгук(для авторизованих користувачів);\n'
                                '- Ctrl + I - відкрити вікно з інформацією;\n'
                                '- Ctrl + A - відкрити панель адміністратора(для вчителя);\n'
                                '1.Увійдіть у свій обліковий запис, щоб мати доступ до своїх оцінок;\n'
                                '2.Оберть потрібний місяць та натисніть кнопку "Оновити";\n'
                                '3.Для зручності Ви можете виділити певний діапазон клітинок, натиснувши '
                                'на заголовки рядків чи стовпців, або виділивши потрібні клітинки. Якщо під '
                                'час виділення тримати клавішу Ctrl, то можна здійснити виділення в кількох окремих '
                                'областях;\n'
                                '4.Авторизовані користувачі можуть надіслати відгуки, натиснувши на кнопку "Відгук";\n')

        self.infoLabel1.setStyleSheet('font-size: 15px')
        self.infoLabel1.setWordWrap(True)
        self.infoLabel2.setStyleSheet('font-size: 15px')
        self.infoLabel2.setWordWrap(True)
        self.mainLayout.addStretch(5)
        self.mainLayout.addWidget(self.infoLabel1)
        self.mainLayout.addWidget(self.linkLabel)
        self.mainLayout.addWidget(self.infoLabel2)

    def closing(self):
        self.close()


class AdminInfoWindow(QWidget):
    def __init__(self):
        QWidget.__init__(self)
        # Window.
        self.setFixedSize(500, 550)
        self.setWindowTitle('Інформація для адміністратора')

        # Layout.
        self.mainLayout = QVBoxLayout(self)

        # Main info label.
        self.infoLabel = QLabel('---Інформація для адміністратора---\n'
                                '1. Відкрийте папку "files", у якій розташовані усі файли. Для введення нового '
                                'класу у цій папці необхідно створити книгу Excel. Для цього відкрийте '
                                'шаблон "example.xlsx" та скопіюйте його вміст до таблиць "table10.xlsx" та '
                                '"table11.xlsx". Для редагування прізвища чи імені учня потрібно змінити їх в '
                                'таблиці Excel та у базі даних Access. \n'
                                '2.Для заповнення бази даних Access під назвою "data.accdb" потрібно '
                                'відкрити таблицю "passwords" та ввести до неї дані про нових учнів, а саме: '
                                'пароль, прізвище, ім\'я, доступ, клас. Учні мають доступ - 1, а адміністратор - 2. \n'
                                '3.Для того, щоб автоматично заповнити усі таблиці Excel прізвищами, '
                                'іменами з бази даних, запустіть програму "parsing.exe", що міститься у корневій '
                                'папці. Для відображення прізвищ у алфавітному порядку відсортуйте '
                                'перший стовпчик аркуша. Для створення аркуша нового предмету скопіюйте '
                                'дані з прізвищами та іменами учнів із відсортованого аркуша. \n'
                                '4.Ви можете через панель адміністратора відкрити таблиці з оцінками 10-го '
                                'та 11-го класів, базу даних, історію дій із програмою, відгуки або запустити '
                                'програму "parsing.exe". \n'
                                'Якщо виникнуть запитання щодо роботи з журналом скористайтеся зворотним зв’язком.')
        self.infoLabel.setStyleSheet('font-size: 15px')
        self.infoLabel.setWordWrap(True)
        self.mainLayout.addWidget(self.infoLabel)


# Support window.
class SupportWindow(QWidget):
    def __init__(self):
        QWidget.__init__(self)
        # Stylesheet.
        try:
            self.setStyleSheet(open('files/styles.css').read())
        except:
            pass

        # Window.
        self.setFixedSize(400, 400)
        self.setWindowTitle('Допоможіть нам покращити програми')

        # Layout.
        self.mainLayout = QVBoxLayout(self)

        # Label.
        self.whatToSendLabel = QLabel('Якщо Ви знайшли помилку в правописі, програмі і т.д., '
                                      'будь ласка, повідомте нас, ми дякуємо за вашу підтримку:')
        self.whatToSendLabel.setStyleSheet('font-size: 15px')
        self.whatToSendLabel.setWordWrap(True)
        self.mainLayout.addWidget(self.whatToSendLabel)

        # Text area.
        self.textArea = QTextEdit()
        self.textArea.setObjectName('textArea')
        self.mainLayout.addWidget(self.textArea)

        # Buttons.
        self.sendButton = QPushButton('Відправити')
        self.sendButton.setObjectName('sendButton')
        self.sendButton.clicked.connect(self.send_info)
        self.sendButton.setShortcut('Return')
        self.mainLayout.addWidget(self.sendButton)

    def send_info(self):
        if self.textArea.toPlainText() == '':
            emptyText = QMessageBox(self)
            emptyText.setWindowTitle('Пусте поле')
            emptyText.setText('Текстове поле не може бути пустим')
            emptyText.setIcon(QMessageBox.Critical)
            emptyText.setStandardButtons(QMessageBox.Ok)
            emptyText.show()
            return

        # Txt file.
        self.txtSupportFile = open('files/feedback.txt', 'a')
        self.txtSupportFile.write(str(datetime.datetime.now().day) + '/' +
                                  str(datetime.datetime.now().month) + '/' +
                                  str(datetime.datetime.now().year) + ' o ' +
                                  str(datetime.datetime.now().hour) + ':' +
                                  str(datetime.datetime.now().minute) + '  ' +
                                  window.nameOfUser +
                                  ' повідомив(ла), що \n' +
                                  self.textArea.toPlainText() + '\n')
        self.txtSupportFile.close()
        self.textArea.clear()
        self.close()


class AdminWindow(QWidget):
    def __init__(self):
        QWidget.__init__(self)

        try:
            self.setStyleSheet(open('files/styles.css').read())
        except:
            pass

        self.setWindowTitle('Панель адміністратора')
        self.setFixedSize(400, 400)

        # Dialogs.
        self.adminInfoDialog = AdminInfoWindow()

        # Layouts.
        self.mainLayout = QGridLayout(self)

        # Buttons.
        self.buttonOpenExcel = QPushButton('Відкрити Excel')
        self.buttonOpenAccess = QPushButton('Відкрити базу даних Access')
        self.buttonOpenLogs = QPushButton('Відкрити історію')
        self.buttonOpenFeedback = QPushButton('Відкрити відгуки')
        self.buttonInfoForAdmin = QPushButton('Інформація для адміністратора')
        self.buttonToParse = QPushButton('Оновити таблицю Excel')
        self.buttonOpenExcel.clicked.connect(self.openExcel)
        self.buttonOpenAccess.clicked.connect(self.openAccess)
        self.buttonOpenLogs.clicked.connect(self.openLogs)
        self.buttonOpenFeedback.clicked.connect(self.openFeedback)
        self.buttonInfoForAdmin.clicked.connect(self.showInfoForAdmin)
        self.buttonToParse.clicked.connect(self.startParsing)
        self.buttonOpenExcel.setObjectName('buttonOpenExcel')
        self.buttonOpenFeedback.setObjectName('buttonOpenFeedback')
        self.buttonOpenLogs.setObjectName('buttonOpenLogs')
        self.buttonOpenAccess.setObjectName('buttonOpenAccess')
        self.buttonInfoForAdmin.setObjectName('buttonInfoForAdmin')
        self.buttonToParse.setObjectName('buttonToParse')

        # Combo boxes.
        self.comboboxOfClasses = QComboBox()
        self.comboboxOfClasses.addItem('10')
        self.comboboxOfClasses.addItem('11')
        self.comboboxOfClasses.setCurrentIndex(-1)

        # Setting layout.
        self.mainLayout.addWidget(self.buttonInfoForAdmin, 1, 0)
        self.mainLayout.addWidget(self.comboboxOfClasses, 0, 0)
        self.mainLayout.addWidget(self.buttonOpenExcel, 0, 1)
        self.mainLayout.addWidget(self.buttonOpenAccess, 1, 1)
        self.mainLayout.addWidget(self.buttonOpenLogs, 2, 1)
        self.mainLayout.addWidget(self.buttonOpenFeedback, 3, 1)
        self.mainLayout.addWidget(self.buttonToParse, 4, 1)

    def showInfoForAdmin(self):
        self.adminInfoDialog.show()

    def startParsing(self):
        pathToParsing = cwd + '/parsing.exe'
        os.startfile(pathToParsing)

    def openExcel(self):
        currentClass = self.comboboxOfClasses.currentText()
        if not currentClass in ('10', '11'):
            errorDialogOfChoosingClass = QMessageBox(self)
            errorDialogOfChoosingClass.setWindowTitle('Помилка вибору класа')
            errorDialogOfChoosingClass.setText('Щоб продовжити оберіть класс')
            errorDialogOfChoosingClass.setIcon(QMessageBox.Critical)
            errorDialogOfChoosingClass.setStandardButtons(QMessageBox.Ok)
            errorDialogOfChoosingClass.show()
            return
        pathToExcelFile = cwd + f'/files/table{currentClass}.xlsx'
        os.startfile(pathToExcelFile)

    def openAccess(self):
        pathToAccessFile = cwd + '/files/data.accdb'
        os.startfile(pathToAccessFile)

    def openLogs(self):
        pathToLogsFile = cwd + '/files/logs.txt'
        os.startfile(pathToLogsFile)

    def openFeedback(self):
        pathToFeedbackFile = cwd + '/files/feedback.txt'
        os.startfile(pathToFeedbackFile)


#   Main window
class MainWindow(QMainWindow):
    def __init__(self):
        QMainWindow.__init__(self)

        # Stylesheet.
        try:
            self.setStyleSheet(open('files/styles.css').read())
        except:
            pass

        # Window.
        self.setWindowTitle('Електронний журнал ЛЗЗСО I-III ст. №4 v.1.0')
        self.setFixedSize(1100, 600)

        # Variables.
        self.arrayOfSubjects = []
        self.arrayOfStudents = []
        self.arrayOfMonth = ('Вересень', 'Жовтень', 'Листопад',  # 0 1 2
                             'Грудень', 'Січень', 'Лютий',  # 3 4 5
                             'Березень', 'Квітень', 'Травень')  # 6 7 8
        self.arrayOfMonthNumbers = (9, 10, 11,
                                    12, 1, 2,
                                    3, 4, 5)
        self.skippedMonth = (1, 0, 1, 0, 0, 3, 0, 1, 0)
        self.classOfStudent = 0
        self.currentColumn = 1
        self.currentRow = 0
        self.isUser = 0
        self.year = datetime.datetime.now().year
        self.skippedColumns = 0
        self.isClear = True
        self.pathToExcelFile = ''
        self.currentMonth = ''
        self.currentName = ''
        self.nameOfUser = ''

        # Dialogs.
        self.supportDialog = SupportWindow()
        self.adminDialog = AdminWindow()
        self.infoDialog = InfoWindow()

        # Layouts.
        self.mainWidget = QWidget()
        self.rightWidget = QWidget()
        self.leftWidget = QWidget()

        self.mainLayout = QHBoxLayout()
        self.rightLayout = QVBoxLayout()
        self.leftLayout = QVBoxLayout()

        self.leftLayoutUp = QBoxLayout(QBoxLayout.TopToBottom)
        self.leftLayoutDown = QBoxLayout(QBoxLayout.TopToBottom)
        self.leftLayout.addLayout(self.leftLayoutUp)
        self.leftLayout.addLayout(self.leftLayoutDown)
        self.mainLayout.addWidget(self.leftWidget)
        self.mainLayout.addWidget(self.rightWidget)

        self.mainWidget.setLayout(self.mainLayout)
        self.leftWidget.setLayout(self.leftLayout)
        self.rightWidget.setLayout(self.rightLayout)
        self.setCentralWidget(self.mainWidget)

        # Buttons.
        self.buttonRefresh = QPushButton('Оновити')
        self.buttonRefresh.setObjectName('buttonRefresh')
        self.buttonInfo = QPushButton('Інформація')
        self.buttonInfo.setObjectName('buttonInfo')
        self.buttonFeedback = QPushButton('Відгук')
        self.buttonFeedback.setObjectName('buttonFeedback')
        self.buttonLogin = QPushButton('Увійти')
        self.buttonLogin.setObjectName('buttonLogin')
        self.buttonExit = QPushButton('Вийти')
        self.buttonExit.setObjectName('buttonExit')
        self.buttonAdmin = QPushButton('Для адміністратора')
        self.buttonAdmin.setObjectName('buttonAdmin')
        self.buttonLogin.clicked.connect(self.enter_the_account)
        self.buttonExit.clicked.connect(self.exit_the_account)
        self.buttonFeedback.clicked.connect(self.feed_back)
        self.buttonRefresh.clicked.connect(self.refresh)
        self.buttonAdmin.clicked.connect(self.for_admin)
        self.buttonInfo.clicked.connect(self.info)

        # Button tips.
        self.buttonFeedback.setToolTip('Ctrl + F')
        self.buttonRefresh.setToolTip('Ctrl + R')
        self.buttonAdmin.setToolTip('Ctrl + A')
        self.buttonExit.setToolTip('Ctrl + Q')
        self.buttonInfo.setToolTip('Ctrl + I')
        self.buttonLogin.setToolTip('Enter')

        # Shortcuts.
        self.buttonFeedback.setShortcut('ctrl+F')
        self.buttonRefresh.setShortcut('ctrl+R')
        self.buttonLogin.setShortcut('Return')
        self.buttonAdmin.setShortcut('ctrl+A')
        self.buttonInfo.setShortcut('ctrl+I')
        self.buttonExit.setShortcut('ctrl+Q')

        # ComboBoxes.
        self.comboBoxOfMonth = QComboBox()
        self.comboBoxOfClasses = QComboBox()
        self.comboBoxOfNames = QComboBox()
        self.comboBoxOfMonth.addItems(self.arrayOfMonth)
        self.comboBoxOfClasses.addItems(['10', '11'])
        self.comboBoxOfClasses.setCurrentIndex(-1)

        # Creator: Shevtsov(Nimchenko) Maksym, 2019-2020 - 11 form.

        self.comboBoxOfClasses.currentIndexChanged.connect(self.select_class)
        self.comboBoxOfNames.setCurrentIndex(-1)

        # Line edit.
        self.enterPass = QLineEdit()
        self.enterPass.setObjectName('enterPassEdit')
        self.enterPass.setEchoMode(QLineEdit.Password)

        # Labels.
        self.greetingUser = QLabel()
        self.labelOfImage = QLabel()
        self.labelOfImage.setStyleSheet(
            'border:10px solid rgb(240, 240, 240);')
        self.labelOfImage.resize(100, 100)

        # User image.
        self.pixmapAdminUser = QPixmap(cwd + '/images/AdminUser.png')
        self.pixmapYesUser = QPixmap(cwd + '/images/YesUser.png')
        self.pixmapNoUser = QPixmap(cwd + '/images/NoUser.png')
        self.labelOfImage.setPixmap(self.pixmapNoUser)

        # Setting layouts.
        self.leftLayoutUp.addWidget(self.labelOfImage)
        self.leftLayoutUp.addWidget(self.enterPass)
        self.leftLayoutUp.addWidget(self.buttonLogin)
        self.leftLayoutUp.addWidget(self.buttonExit)
        self.leftLayoutUp.addWidget(self.greetingUser)

        self.leftLayoutDown.addWidget(self.comboBoxOfMonth)
        self.leftLayoutDown.addWidget(self.buttonRefresh)
        self.leftLayoutDown.addWidget(self.buttonInfo)

        self.mainLayout.addStretch(1)
        self.leftLayout.addStretch(1)
        self.leftLayoutDown.setContentsMargins(0, 50, 0, 0)

        # Table widget.
        self.table = QTableWidget()
        self.rightLayout.addWidget(self.table)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setFixedWidth(900)
        self.table.horizontalHeader().setSectionResizeMode(3)
        self.table.verticalHeader().setSectionResizeMode(1)

        self.detectExistingFiles()

    def check(self):
        print('check start')
        print('check end')

    def feed_back(self):
        self.supportDialog.show()

    def info(self):
        self.infoDialog.show()

    def for_admin(self):
        self.adminDialog.show()
        logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
        logTextFile.write(str(datetime.datetime.now().day) + '/' +
                          str(datetime.datetime.now().month) + '/' +
                          str(datetime.datetime.now().year) + ' o ' +
                          str(datetime.datetime.now().hour) + ':' +
                          str(datetime.datetime.now().minute) + '  ' +
                          'Панель адміністратора відкрита' + '\n')
        logTextFile.close()

    def detectExistingFiles(self):
        # images: adminUser.png, icon.ico, noUser.png, parseIcon.ico, yesUser.png
        # files: data.accdb, feedback.txt, logs.txt, styles.css, table10.xlsx, table11.xlsx
        files = os.path.exists('files')
        images = os.path.exists('images')

        # Images.
        adminUserPic = os.path.exists('images/adminUser.png')
        yesUserPic = os.path.exists('images/yesUser.png')
        noUserPic = os.path.exists('images/noUser.png')
        parseIcon = os.path.exists('images/parseIcon.ico')
        appIcon = os.path.exists('images/icon.png')
        arrayOfImages = [images, 'images',
                         adminUserPic, 'images/adminUser.png',
                         yesUserPic, 'images/yesUser.png',
                         noUserPic, 'images/noUser.png',
                         parseIcon, 'images/parseIcon.ico',
                         appIcon, 'images/icon.png']

        # Files.
        data = os.path.exists('files/data.accdb')
        feedback = os.path.exists('files/feedback.txt')
        logs = os.path.exists('files/logs.txt')
        styles = os.path.exists('files/styles.css')
        table10 = os.path.exists('files/table10.xlsx')
        table11 = os.path.exists('files/table11.xlsx')
        arrayOfFiles = [files, 'files',
                        data, 'files/data.accdb',
                        feedback, 'files/feedback.txt',
                        logs, 'files/logs.txt',
                        styles, 'files/styles.css',
                        table10, 'files/table10.xlsx',
                        table11, 'files/table11.xlsx']

        arrayOfNonExistingFiles = []
        for i in range(len(arrayOfFiles)):
            if arrayOfFiles[i] == False:
                arrayOfNonExistingFiles.append(arrayOfFiles[i+1] + '\n')
        for i in range(len(arrayOfImages)):
            if arrayOfImages[i] == False:
                arrayOfNonExistingFiles.append(arrayOfImages[i+1] + '\n')

        # Error dialog.
        if arrayOfNonExistingFiles:
            errorEnterDialog = QMessageBox(self)
            errorEnterDialog.setWindowTitle('Помилка завантаження файлів')
            errorEnterDialog.setText(
                'Програма може працювати неправильно, бо бракує таких файлів як:\n' + ''.join(arrayOfNonExistingFiles))
            errorEnterDialog.setIcon(QMessageBox.Warning)
            errorEnterDialog.setStandardButtons(QMessageBox.Ok)
            errorEnterDialog.show()

    def select_class(self):
        if self.comboBoxOfClasses.currentIndex() == 0:
            currentClass = 10
        elif self.comboBoxOfClasses.currentIndex() == 1:
            currentClass = 11

        self.pathToExcelFile = cwd + f'/files/table{currentClass}.xlsx'
        self.excelFile = openpyxl.load_workbook(
            filename=self.pathToExcelFile, read_only=True, keep_vba=False)
        self.arrayOfSubjects = self.excelFile.sheetnames
        self.excelFile.close()
        self.arrayOfStudents.clear()
        self.comboBoxOfNames.clear()

        for row in cursor.execute("select * from passwords"):
            if row.room == currentClass:
                self.arrayOfStudents.append(row.names)
        self.classOfStudent = currentClass
        self.arrayOfStudents.sort()
        self.comboBoxOfNames.addItems(self.arrayOfStudents)

    def refresh(self):
        if self.isUser == 0:
            errorEnterDialog = QMessageBox(self)
            errorEnterDialog.setWindowTitle('Помилка оновлення')
            errorEnterDialog.setText(
                'Для того, щоб переглянути свої досягнення, увійдіть у свій обліковий запис')
            errorEnterDialog.setIcon(QMessageBox.Critical)
            errorEnterDialog.setStandardButtons(QMessageBox.Ok)
            errorEnterDialog.show()
            return
        if self.isUser == 2:
            if self.comboBoxOfClasses.currentText() == '' or self.comboBoxOfNames.currentText() == '':
                errorEnterDialog = QMessageBox(self)
                errorEnterDialog.setWindowTitle('Помилка оновлення')
                errorEnterDialog.setText('Для того, щоб переглянути досягнення учня(учениці), '
                                         'оберіть у списку необхідні дані')
                errorEnterDialog.setIcon(QMessageBox.Critical)
                errorEnterDialog.setStandardButtons(QMessageBox.Ok)
                errorEnterDialog.show()
                return
        self.year = datetime.datetime.now().year
        self.table.setColumnCount(0)
        self.table.setRowCount(0)
        self.currentMonth = self.comboBoxOfMonth.currentText()
        if self.isUser != 2:
            self.currentName = self.nameOfUser
        elif self.isUser == 2:
            self.currentName = self.comboBoxOfNames.currentText()
        self.isClear = False
        isIntercalary = False

        # Previous year.
        curMonth = datetime.datetime.now().month
        if curMonth in (9, 10, 11, 12) and \
                not self.arrayOfMonthNumbers[self.arrayOfMonth.index(self.currentMonth)] in (9, 10, 11, 12):
            self.year += 1
        if curMonth in (1, 2, 3, 4, 5) and \
                not self.arrayOfMonthNumbers[self.arrayOfMonth.index(self.currentMonth)] in (1, 2, 3, 4, 5):
            self.year -= 1

        # Check the intercalary year.
        if self.year % 4 != 0 or (self.year % 100 == 0 and self.year % 400 != 0):
            isIntercalary = False
        else:
            isIntercalary = True

        # Updating rows labels.
        self.excelFile = openpyxl.load_workbook(
            filename=self.pathToExcelFile, read_only=True, keep_vba=False)
        self.arrayOfSubjects = self.excelFile.sheetnames
        self.excelFile.close()

        # Setting table.
        self.table.setRowCount(len(self.arrayOfSubjects))
        self.table.setVerticalHeaderLabels(self.arrayOfSubjects)

        # Parsing data to from Excel to QTableWidget.
        book = openpyxl.load_workbook(self.pathToExcelFile, data_only=True)
        startOfMonth = 0
        endOfMonth = 0
        masArrayOfDays = []
        masArrayOfMarks = []

        # Searching for number of days and average.
        for i in range(len(self.arrayOfSubjects)):
            currentSheet = book[self.arrayOfSubjects[i]]
            for j in range(1, currentSheet.max_row + 1):
                if self.currentName == currentSheet.cell(row=j, column=1).value:
                    for a in range(1, 500):
                        cell = currentSheet.cell(row=1, column=a).value
                        if self.currentMonth == cell:
                            startOfMonth = a
                        if str(cell) == cell:
                            if self.arrayOfMonth.index(cell) - self.arrayOfMonth.index(self.currentMonth) == 1:
                                endOfMonth = a - 1
                        if self.currentMonth == 'Травень' and \
                                not (isinstance(currentSheet.cell(row=2, column=a).value, int) or isinstance(
                                    currentSheet.cell(row=2, column=a).value, str)):
                            endOfMonth = a - 1
                            break
                    arrayOfDays = []
                    arrayOfMarks = []
                    for b in range(startOfMonth, endOfMonth + 1):
                        arrayOfMarks.append(
                            currentSheet.cell(row=j, column=b).value)
                        arrayOfDays.append(
                            currentSheet.cell(row=2, column=b).value)
                    masArrayOfDays.append(arrayOfDays)
                    masArrayOfMarks.append(arrayOfMarks)

        # Setting days into array.
        array = []
        for i in masArrayOfDays[0]:
            if isinstance(i, int):
                array.append(i)
        for ar in masArrayOfDays:
            for a in range(1, len(ar)):
                if ar[0] == 'т' and array[0] != 'т':
                    array.insert(0, 'т')
                if ar[len(ar) - 1] == 'т' and array[len(array) - 1] != 'т':
                    array.insert(len(array), 'т')
                if ar[a] == 'т':
                    prev = ar[a - 1]
                    for j in range(0, len(array) - 1):
                        if array[j] == prev and array[j + 1] != 'т':
                            array.insert(j + 1, 'т')

        # Setting marks.
        for i in range(len(masArrayOfDays)):
            for j in range(len(masArrayOfDays[i])):
                if isinstance(masArrayOfDays[i][j], str):
                    masArrayOfMarks[i][j] = str(masArrayOfMarks[i][j]) + 't'
        for i in range(len(masArrayOfMarks)):
            for a in range(len(masArrayOfMarks[i])):
                for j in range(len(array)):
                    if isinstance(array[j], str):
                        if isinstance(masArrayOfMarks[i][j], int):
                            masArrayOfMarks[i].insert(j, '')

        # Make all string.
        for i in range(len(array)):
            array[i] = str(array[i])

        # Delete weekends.
        for i in range(len(array)):
            try:
                temp = int(array[i])
            except:
                continue
            if temp == 29 and self.currentMonth == 'Лютий' and not isIntercalary:
                continue
            day = datetime.date(self.year,
                                self.arrayOfMonthNumbers[self.arrayOfMonth.index(
                                    self.currentMonth)],
                                temp).isoweekday()

            if day == 1:
                array[i] += ' Пн.'
            if day == 2:
                array[i] += ' Вт.'
            if day == 3:
                array[i] += ' Ср.'
            if day == 4:
                array[i] += ' Чт.'
            if day == 5:
                array[i] += ' Пт.'
            if day == 6:
                array[i] += ' Сб.'
            if day == 7:
                array[i] += ' Нд.'
        self.table.setColumnCount(len(array))
        self.table.setHorizontalHeaderLabels(array)

        # Delete 29th in not Intencelary year.
        if not isIntercalary and self.currentMonth == 'Лютий':
            for i in range(len(array)):
                if int(array[i][:2]) == 29:
                    self.table.setColumnHidden(i, True)

        # Deleting weekends days.
        for i in range(len(array)):
            temp = array[i]
            if temp[-3:] == 'Сб.' or temp[-3:] == 'Нд.':
                self.table.setColumnHidden(i, True)

        # Parsing to table.
        for i in range(len(masArrayOfMarks)):
            for j in range(len(array)):
                cell = masArrayOfMarks[i][j]
                cell = str(cell)
                cell = cell.replace('t', '')
                if cell == 'None':
                    cell = ''
                if '.' in cell:
                    cell = float(cell)
                    cell = str(round(cell, 2))
                self.table.setItem(i, j, QTableWidgetItem(str(cell)))
                try:
                    cell = float(cell)
                    r = cell * 20
                    g = cell * 20
                    if cell > 8:
                        self.table.item(i, j).setBackground(
                            QColor(0, g, 0, 200))
                    elif cell < 9:
                        self.table.item(i, j).setBackground(
                            QColor(255 - r, g, 0, 200))
                except:
                    if str(cell) == 'н':
                        self.table.item(i, j).setBackground(
                            QColor(255, 255, 0, 200))

        logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
        tempText = ''
        if self.isUser == 2:
            tempText = f", клас: {self.classOfStudent}, ім'я: {self.currentName}"
        if self.isUser == 2:
            tempClass = 'admin'
        else:
            tempClass = self.classOfStudent
        logTextFile.write(str(datetime.datetime.now().day) + '/' +
                          str(datetime.datetime.now().month) + '/' +
                          str(datetime.datetime.now().year) + ' o ' +
                          str(datetime.datetime.now().hour) + ':' +
                          str(datetime.datetime.now().minute) + '  ' +
                          self.nameOfUser +
                          f'(клас: {tempClass}) оновив(ла) сторінку з даними в місяці: '
                          + self.currentMonth + tempText + '\n')
        logTextFile.close()

    def enter_the_account(self):
        countOfSuccess = 0
        stringOfPass = self.enterPass.text()

        # Check whether app can go on.
        if self.isUser != 0 and stringOfPass != '':
            logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
            logTextFile.write(str(datetime.datetime.now().day) + '/' +
                              str(datetime.datetime.now().month) + '/' +
                              str(datetime.datetime.now().year) + ' o ' +
                              str(datetime.datetime.now().hour) + ':' +
                              str(datetime.datetime.now().minute) + '  ' +
                              'Вхід в інший аккаунт з використанням паролю ' + stringOfPass + '\n')
            logTextFile.close()
        if self.isUser == 0 and stringOfPass == '' or self.isUser != 0:
            return

        # If user exit account so clear the table.
        if self.isClear == False:
            self.clear_the_table()
            self.isClear = True

        # Parsing data to app from Access Data Base.
        for row in cursor.execute("select * from passwords"):
            if stringOfPass == row.passes:
                countOfSuccess += 1
                self.isUser = row.access
                self.nameOfUser = row.names
                self.classOfStudent = row.room
                tempOfTimeOfDay = ''
                if datetime.datetime.now().hour in (7, 8, 9, 10, 11):
                    tempOfTimeOfDay = 'Добрий ранок, '
                elif datetime.datetime.now().hour in (12, 13, 14, 15, 16, 17):
                    tempOfTimeOfDay = 'Добрий день, '
                elif datetime.datetime.now().hour in (18, 19, 20, 21):
                    tempOfTimeOfDay = 'Добрий вечір, '
                elif datetime.datetime.now().hour in (22, 23, 24, 1, 2, 3, 4, 5):
                    tempOfTimeOfDay = 'Чому ви не спите?'
                self.greetingUser.setText(tempOfTimeOfDay + self.nameOfUser)
                self.greetingUser.setWordWrap(True)
                break

        # Wrong password.
        if countOfSuccess == 0 and stringOfPass != '':
            errorEnterDialog = QMessageBox(self)
            errorEnterDialog.setWindowTitle('Помилка входу')
            errorEnterDialog.setText(
                'Невірний пароль.Будь ласка, перевірте правильність паролю')
            errorEnterDialog.setIcon(QMessageBox.Critical)
            errorEnterDialog.setStandardButtons(QMessageBox.Ok)
            errorEnterDialog.show()
            logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
            logTextFile.write(str(datetime.datetime.now().day) + '/' +
                              str(datetime.datetime.now().month) + '/' +
                              str(datetime.datetime.now().year) + ' o ' +
                              str(datetime.datetime.now().hour) + ':' +
                              str(datetime.datetime.now().minute) + '  ' +
                              'Невірний пароль ' + stringOfPass + '\n')
            logTextFile.close()
            return

        # Setting an image of user.
        if countOfSuccess == 1:
            self.enterPass.setText('')
            self.leftLayoutDown.addWidget(self.buttonFeedback)
            if self.isUser == 2:
                tempClass = 'admin'
            else:
                tempClass = self.classOfStudent
            logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
            logTextFile.write(str(datetime.datetime.now().day) + '/' +
                              str(datetime.datetime.now().month) + '/' +
                              str(datetime.datetime.now().year) + ' o ' +
                              str(datetime.datetime.now().hour) + ':' +
                              str(datetime.datetime.now().minute) + '  ' +
                              'Користувач ' + self.nameOfUser + f'(клас: {tempClass})' + ' увійшов(ла) в аккаунт\n')
            logTextFile.close()

        # Loading sheets.
        if self.isUser != 2:
            self.pathToExcelFile = cwd + \
                f'/files/table{self.classOfStudent}.xlsx'
            self.labelOfImage.setPixmap(self.pixmapYesUser)
        if self.isUser == 2:
            self.leftLayoutDown.insertWidget(0, self.comboBoxOfNames)
            self.leftLayoutDown.insertWidget(0, self.comboBoxOfClasses)
            self.leftLayoutDown.addWidget(self.buttonAdmin)
            self.labelOfImage.setPixmap(self.pixmapAdminUser)

    def clear_the_table(self):
        self.table.setRowCount(0)
        self.table.setColumnCount(0)

    def exit_the_account(self):
        if self.isUser == 0:
            return
        if self.isUser == 2:
            tempClass = 'admin'
        else:
            tempClass = self.classOfStudent
        logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
        logTextFile.write(str(datetime.datetime.now().day) + '/' +
                          str(datetime.datetime.now().month) + '/' +
                          str(datetime.datetime.now().year) + ' o ' +
                          str(datetime.datetime.now().hour) + ':' +
                          str(datetime.datetime.now().minute) + '  ' +
                          'Користувач ' + self.nameOfUser + f'(клас: {tempClass})' + ' вийшов з аккаунту\n')
        logTextFile.close()
        if self.isUser == 2:
            self.comboBoxOfNames.setParent(None)
            self.comboBoxOfClasses.setParent(None)
            self.buttonAdmin.setParent(None)
        self.labelOfImage.setPixmap(self.pixmapNoUser)
        self.buttonFeedback.setParent(None)
        self.greetingUser.setText('')
        self.isUser = 0
        if self.isClear == False:
            self.clear_the_table()
            self.table.setRowCount(0)
            self.table.setColumnCount(0)

    def closeEvent(self, *args, **kwargs):
        if self.isUser not in (1, 2):
            return
        if self.isUser == 2:
            tempClass = 'admin'
        else:
            tempClass = self.classOfStudent
        if self.nameOfUser == '':
            tempText = 'Невідомий користувач закрив(ла) програму'
        else:
            tempText = 'Користувач ' + self.nameOfUser + f'(клас: {tempClass})' + \
                       ' вийшов(ла) з аккаунту(закрив програму)'
        logTextFile = open('files/logs.txt', 'a', encoding='utf-8')
        logTextFile.write(str(datetime.datetime.now().day) + '/' +
                          str(datetime.datetime.now().month) + '/' +
                          str(datetime.datetime.now().year) + ' o ' +
                          str(datetime.datetime.now().hour) + ':' +
                          str(datetime.datetime.now().minute) + '  ' +
                          tempText + '\n')
        logTextFile.close()


if __name__ == '__main__':
    app = QApplication(sys.argv)

    app.setStyle('Fusion')
    appIcon = QIcon()
    appIcon.addFile(cwd + '/images/icon.png', QSize(64, 64))
    app.setWindowIcon(appIcon)
    window = MainWindow()
    window.show()

    sys.exit(app.exec_())  # Creator: Shevtsov(Nimchenko) Maksym, 2019-2020 - 11 form.
